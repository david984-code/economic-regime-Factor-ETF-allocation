import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# SETTINGS 
INPUT_CSV = "optimal_allocations.csv"
OUTPUT_XLSX = "optimal_allocations_formatted.xlsx"

THRESHOLD = 0.0001   # 0.01% cutoff (set tiny weights to 0)
PCT_DECIMALS = 2     # show 2 decimal places in percentage

# Read allocations
df = pd.read_csv(INPUT_CSV)

# Expect first column is regime 
if "regime" in df.columns:
    df = df.set_index("regime")
else:
    df = df.set_index(df.columns[0])

# 1) Zero out tiny weights
df = df.astype(float)
df[np.abs(df) < THRESHOLD] = 0.0

# 2) Renormalize each row to sum to 1.0
row_sums = df.sum(axis=1)
for idx, s in row_sums.items():
    if s > 0:
        df.loc[idx, :] = df.loc[idx, :] / s

# 3) Optional: round weights (in weight units), then renormalize again
df = df.round(6)
row_sums = df.sum(axis=1)
for idx, s in row_sums.items():
    if s > 0:
        df.loc[idx, :] = df.loc[idx, :] / s

# Convert to % for display (still stored as decimals in Excel formatting below)
df_out = df.copy()
df_out.insert(0, "regime", df_out.index)

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "allocations"

# Write data
for r in dataframe_to_rows(df_out, index=False, header=True):
    ws.append(r)

# Formatting
bold = Font(bold=True)

# Bold header row
for cell in ws[1]:
    cell.font = bold

# Bold first column)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.font = bold
        cell.alignment = Alignment(horizontal="left")

# Set percent formatting for numeric cells 
percent_format = "0." + ("0" * PCT_DECIMALS) + "%"
for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, (int, float, np.floating)):
            cell.number_format = percent_format

# Autosize columns a bit
for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        val = "" if cell.value is None else str(cell.value)
        max_len = max(max_len, len(val))
    ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

wb.save(OUTPUT_XLSX)
print(f"✅ Saved formatted Excel file: {OUTPUT_XLSX}")
