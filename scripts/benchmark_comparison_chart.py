"""Compare frozen baseline strategy vs index proxies over 1Y / 5Y / 10Y / 15Y / 20Y windows.

Benchmarks (ETF proxies, total-return via adjusted close):
  SPY ~ S&P 500 (SPX), QQQ ~ Nasdaq-100 (NDX), IWM ~ Russell 2000 (RTY).

Strategy: one full-sample backtest (optimizer allocations trained on full monthly history,
then run_backtest_with_allocations with frozen baseline knobs — same spirit as walk-forward
experiments, but NOT walk-forward OOS). See printed disclaimer.

Outputs:
  outputs/benchmark_comparison_metrics.xlsx — **open this in Excel** (native % / Sharpe, colors, column widths)
  outputs/benchmark_comparison_metrics.csv — raw floats for code; opening in Excel shows decimals (CSV has no Excel styling)
  outputs/benchmark_comparison_metrics_formatted.csv — text like "16.84%" (still no colors if opened as CSV)
  outputs/benchmark_comparison_chart.png  (requires matplotlib: uv add matplotlib)

Usage (repo root):
  uv run python scripts/benchmark_comparison_chart.py
"""

# This script intentionally modifies sys.path before importing project modules.
# Ruff would otherwise flag E402 (imports not at top of file).
# ruff: noqa: E402

from __future__ import annotations

import sys
import warnings
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
warnings.filterwarnings("ignore")

import csv
import logging
import math
from datetime import datetime

import pandas as pd
import yfinance as yf

from src.allocation.optimizer import optimize_allocations_from_data
from src.backtest.engine import run_backtest_with_allocations
from src.backtest.metrics import compute_metrics
from src.config import OUTPUTS_DIR, TICKERS, get_end_date
from src.data.market_ingestion import fetch_prices
from src.evaluation.benchmarks import CASH_DAILY_YIELD
from src.utils.retry import retry_on_permission_error

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# Frozen baseline (aligned with accepted experiments; inv-vol OFF post-blend)
BASELINE_KW = {
    "use_stagflation_override": False,
    "use_hybrid_signal": True,
    "hybrid_macro_weight": 0.0,
    "use_momentum": True,
    "market_lookback_months": 24,
    "trend_filter_type": "none",
    "vol_scaling_method": "none",
    "portfolio_construction_method": "equal_weight",
    "momentum_12m_weight": 0.0,
    "quarterly_rebalance": False,
    "tolerance": 0.015,
    "sigmoid_scale": 0.25,
    "use_post_blend_inv_vol": False,
}

BENCHMARK_TICKERS = {
    "SPY": "S&P 500 (SPY~SPX)",
    "QQQ": "Nasdaq-100 (QQQ~NDX)",
    "IWM": "Russell 2000 (IWM~RTY)",
}
HORIZONS_YEARS = (1, 5, 10, 15, 20)


def _load_regime_df(prices: pd.DataFrame) -> pd.DataFrame:
    csv_path = OUTPUTS_DIR / "regime_labels_expanded.csv"
    if csv_path.exists():
        regime_df = pd.read_csv(csv_path, parse_dates=["date"], index_col="date")
    else:
        from src.allocation.optimizer import load_regimes

        regime_df = load_regimes()
    regime_df = regime_df.dropna(subset=["regime"]).sort_index()
    if regime_df.index.duplicated().any():
        regime_df = regime_df[~regime_df.index.duplicated(keep="last")]
    regime_df = regime_df.reindex(prices.index).ffill()
    return regime_df


def _run_strategy_returns() -> pd.Series:
    end = get_end_date()
    prices = fetch_prices(tickers=TICKERS, start="2010-01-01", end=end)
    regime_df = _load_regime_df(prices)

    train_returns = prices.resample("ME").last().pct_change().dropna()
    if "cash" not in train_returns.columns:
        train_returns["cash"] = (1.05) ** (1 / 12) - 1
    train_regimes = (
        regime_df.loc[: train_returns.index.max()].resample("ME").last().dropna(how="all")
    )
    train_regimes = train_regimes.loc[train_regimes.index <= train_returns.index.max()]

    allocations = optimize_allocations_from_data(train_returns, train_regimes)
    if not allocations:
        raise RuntimeError("optimize_allocations_from_data returned empty allocations")

    result = run_backtest_with_allocations(
        prices,
        regime_df,
        allocations,
        return_weights=False,
        **BASELINE_KW,
    )
    if isinstance(result, tuple):
        strat_rets = result[0]
    else:
        strat_rets = result
    return strat_rets.dropna()


def _download_benchmarks(end: str) -> pd.DataFrame:
    """Daily total-return proxy: adjusted close -> pct_change."""
    tickers = list(BENCHMARK_TICKERS.keys())
    raw = yf.download(
        tickers, start="1998-01-01", end=end, progress=False, auto_adjust=True, threads=True
    )
    if raw.empty:
        raise RuntimeError("yfinance returned no benchmark data")
    close = raw["Close"] if "Close" in raw.columns else raw
    if isinstance(close, pd.Series):
        close = close.to_frame(name=tickers[0])
    rets = close.pct_change().iloc[1:].dropna(how="all")
    return rets


def _slice_window(series: pd.Series, end_ts: pd.Timestamp, years: int) -> pd.Series:
    start_ts = end_ts - pd.DateOffset(years=years)
    s = series.loc[(series.index >= start_ts) & (series.index <= end_ts)].dropna()
    return s


def _excel_fill_for_series(series_name: str):
    """Row background by series so strategy vs benchmarks scan quickly in Excel."""
    from openpyxl.styles import PatternFill

    # Distinct, print-friendly pastels (strategy highlighted vs three benchmarks)
    fills: dict[str, PatternFill] = {
        "Strategy (frozen baseline)": PatternFill("solid", fgColor="C6E0B4"),
        "S&P 500 (SPY~SPX)": PatternFill("solid", fgColor="BDD7EE"),
        "Nasdaq-100 (QQQ~NDX)": PatternFill("solid", fgColor="F8CBAD"),
        "Russell 2000 (IWM~RTY)": PatternFill("solid", fgColor="FFE699"),
    }
    return fills.get(str(series_name), PatternFill("solid", fgColor="FFFFFF"))


def _style_metrics_excel_sheet(ws, df: pd.DataFrame) -> None:
    """Excel formats: % / Sharpe; row colors by series; header row style."""
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    ncols = len(df.columns)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font

    pct_fmt = "0.00%"
    sharpe_fmt = "0.000"
    n_data_rows = len(df)
    for r in range(2, n_data_rows + 2):
        row_idx = r - 2
        series_name = df.iloc[row_idx]["series"]
        row_fill = _excel_fill_for_series(series_name)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = row_fill
        for c, fmt in ((3, pct_fmt), (4, pct_fmt), (5, sharpe_fmt), (6, pct_fmt)):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            if isinstance(v, float) and math.isnan(v):
                continue
            if not isinstance(v, (int, float)):
                continue
            cell.number_format = fmt
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    for col in "CDEF":
        ws.column_dimensions[col].width = 14


def _write_excel_with_fallback(df: pd.DataFrame, path: Path) -> Path | None:
    """Write .xlsx with native Excel % / Sharpe formats. Returns path or None if openpyxl missing."""
    try:
        __import__("openpyxl")
    except ImportError:
        logger.warning("openpyxl not available; skipping .xlsx (install: uv add openpyxl)")
        return None

    def _write_to(p: Path) -> None:
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="metrics")
            _style_metrics_excel_sheet(writer.sheets["metrics"], df)

    try:
        retry_on_permission_error(
            lambda: _write_to(path), max_attempts=4, base_wait=0.5, logger=logger
        )
        return path
    except PermissionError:
        alt = path.parent / f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        logger.warning(
            "Could not write %s after retries (close Excel). Writing %s instead.",
            path,
            alt.name,
        )
        _write_to(alt)
        return alt


def _write_csv_with_fallback(df: pd.DataFrame, path: Path) -> Path:
    """Write CSV; retry locked files with backoff, then timestamped name."""

    def _do_write(p: Path) -> None:
        df.to_csv(p, index=False, quoting=csv.QUOTE_ALL)

    try:
        retry_on_permission_error(
            lambda: _do_write(path), max_attempts=4, base_wait=0.5, logger=logger
        )
        return path
    except PermissionError:
        alt = path.parent / f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        logger.warning(
            "Could not write %s after retries. Writing %s instead.",
            path,
            alt.name,
        )
        _do_write(alt)
        return alt


def _format_metrics_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Human-readable copy: CAGR / Vol / Max DD as percent strings; Sharpe to 3 decimals."""
    out = df[["horizon", "series"]].copy()
    out["CAGR"] = df["CAGR"].map(lambda x: f"{100 * x:.2f}%" if pd.notna(x) else "")
    out["Volatility"] = df["Volatility"].map(lambda x: f"{100 * x:.2f}%" if pd.notna(x) else "")
    out["Sharpe"] = df["Sharpe"].map(lambda x: f"{x:.3f}" if pd.notna(x) else "")
    out["Max_Drawdown"] = df["Max_Drawdown"].map(lambda x: f"{100 * x:.2f}%" if pd.notna(x) else "")
    return out


def _metrics_row(name: str, rets: pd.Series) -> dict[str, float]:
    m = compute_metrics(rets, rf_daily=CASH_DAILY_YIELD)
    return {
        "name": name,
        "CAGR": m["CAGR"],
        "Volatility": m["Volatility"],
        "Sharpe": m["Sharpe"],
        "Max_Drawdown": m["Max Drawdown"],
    }


def main() -> int:
    print(
        "\nDISCLAIMER: Strategy metrics use a single full-sample backtest (allocations trained once\n"
        "on all monthly history). This is NOT walk-forward out-of-sample. Benchmarks use ETF\n"
        "proxies (SPY/QQQ/IWM), not index total-return indices directly.\n"
    )

    end_str = get_end_date()
    bench_rets = _download_benchmarks(end_str)
    strat_rets = _run_strategy_returns()

    common_end = min(strat_rets.index.max(), bench_rets.index.max())
    strat_start = strat_rets.index.min()
    logger.info("Common end date for slices: %s", common_end.date())
    logger.info(
        "Strategy daily series starts %s (~%.1f years to end); horizons longer than that reuse the same strategy slice",
        strat_start.date(),
        (common_end - strat_start).days / 365.25,
    )

    rows: list[dict] = []
    for years in HORIZONS_YEARS:
        label = f"{years}Y"
        s_strat = _slice_window(strat_rets, common_end, years)
        if len(s_strat) < 60:
            strat_m = {k: float("nan") for k in ("CAGR", "Volatility", "Sharpe", "Max_Drawdown")}
        else:
            strat_m = _metrics_row("Strategy", s_strat)
            strat_m = {k: strat_m[k] for k in ("CAGR", "Volatility", "Sharpe", "Max_Drawdown")}
        rows.append({"horizon": label, "series": "Strategy (frozen baseline)", **strat_m})

        for sym, desc in BENCHMARK_TICKERS.items():
            if sym not in bench_rets.columns:
                continue
            s_b = _slice_window(bench_rets[sym], common_end, years)
            if len(s_b) < 60:
                bench_m = {
                    k: float("nan") for k in ("CAGR", "Volatility", "Sharpe", "Max_Drawdown")
                }
            else:
                bench_m = _metrics_row(sym, s_b)
                bench_m = {k: bench_m[k] for k in ("CAGR", "Volatility", "Sharpe", "Max_Drawdown")}
            rows.append({"horizon": label, "series": desc, **bench_m})

    df = pd.DataFrame(rows)
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    out_csv = OUTPUTS_DIR / "benchmark_comparison_metrics.csv"
    out_fmt = OUTPUTS_DIR / "benchmark_comparison_metrics_formatted.csv"
    fmt_df = _format_metrics_for_display(df)
    # QUOTE_ALL avoids Excel mis-reading columns when text fields contain commas/special chars
    written_csv = _write_csv_with_fallback(df, out_csv)
    written_fmt = _write_csv_with_fallback(fmt_df, out_fmt)
    out_xlsx = OUTPUTS_DIR / "benchmark_comparison_metrics.xlsx"
    written_xlsx = _write_excel_with_fallback(df, out_xlsx)
    logger.info("Wrote %s (numeric)", written_csv)
    logger.info("Wrote %s (CAGR/Vol/MaxDD as %%, Sharpe rounded)", written_fmt)
    if written_xlsx is not None:
        logger.info("Wrote %s (Excel native %% formats)", written_xlsx)
    if (
        written_csv != out_csv
        or written_fmt != out_fmt
        or (written_xlsx is not None and written_xlsx != out_xlsx)
    ):
        print(
            "\nTIP: Close the CSV/XLSX/PNG in Excel if you want to overwrite the default filenames next run.\n"
        )
    abs_csv = written_csv.resolve()
    abs_fmt = written_fmt.resolve()
    if written_xlsx is not None:
        print(
            "\n--- Excel review: use the workbook below (not the .csv) ---\n"
            f"  {written_xlsx.resolve()}\n"
            "  CSV is plain text: double-clicking a .csv in Excel will NOT apply % formats,\n"
            "  row colors, or saved column widths — only the .xlsx file contains those.\n"
        )
    else:
        print("\nExcel workbook not written (install openpyxl for .xlsx).")
    print(f"Numeric CSV (for scripts / pipelines; raw decimals in Excel):\n  {abs_csv}")
    print(f"Formatted CSV (percent as text; still no Excel colors):\n  {abs_fmt}\n")

    # Full table in console (CAGR, Vol, Sharpe, Max DD — same columns as CSV)
    display = df.copy()
    for c in ("CAGR", "Volatility", "Max_Drawdown"):
        display[c] = display[c].map(lambda x: f"{x:7.2%}" if pd.notna(x) else "   n/a")
    display["Sharpe"] = display["Sharpe"].map(lambda x: f"{x:6.3f}" if pd.notna(x) else "   n/a")
    print("All metrics (also in CSV):")
    print(display.to_string(index=False))

    # Pivot for CAGR-only quick view
    pivot = df.pivot_table(index="series", columns="horizon", values="CAGR")
    print("\nCAGR by horizon (rows=series, cols=horizon):")
    print(pivot.to_string(float_format=lambda x: f"{x:7.2%}" if pd.notna(x) else "   n/a"))

    try:
        import matplotlib.pyplot as plt
    except ImportError:
        print("\nInstall matplotlib for PNG chart: uv add matplotlib")
        return 0

    horizons_labels = [f"{y}Y" for y in HORIZONS_YEARS]
    series_order = ["Strategy (frozen baseline)"] + list(BENCHMARK_TICKERS.values())
    # Match Excel row tints: strategy green, SPY blue, Nasdaq coral, Russell gold
    bar_colors = ["#548235", "#2E75B6", "#E97132", "#BF8F00"]
    bar_labels = [ser.split("(")[0].strip() for ser in series_order]

    fig, axes = plt.subplots(2, 2, figsize=(12, 9))
    fig.suptitle(
        "Strategy vs benchmarks (ETF proxies)\n"
        "Strategy = full-sample backtest; benchmarks = SPY/QQQ/IWM adjusted-close returns",
        fontsize=11,
    )
    metrics_config = [
        ("CAGR", "CAGR", axes[0, 0], lambda v: v * 100, "%"),
        ("Volatility", "Ann. volatility", axes[0, 1], lambda v: v * 100, "%"),
        ("Sharpe", "Sharpe (vs ~4.5% cash in metric)", axes[1, 0], lambda v: v, ""),
        ("Max_Drawdown", "Max drawdown", axes[1, 1], lambda v: v * 100, "%"),
    ]

    x = range(len(horizons_labels))
    width = 0.2
    for col, title, ax, scale, suffix in metrics_config:
        for i, ser in enumerate(series_order):
            vals = []
            for h in horizons_labels:
                cell = df[(df["horizon"] == h) & (df["series"] == ser)]
                v = cell[col].iloc[0] if len(cell) else float("nan")
                vals.append(scale(v) if pd.notna(v) else float("nan"))
            offset = (i - 1.5) * width
            is_strategy = i == 0
            ax.bar(
                [xi + offset for xi in x],
                vals,
                width,
                label=bar_labels[i],
                color=bar_colors[i % len(bar_colors)],
                edgecolor=("#1B4D1B" if is_strategy else "#333333"),
                linewidth=1.4 if is_strategy else 0.6,
                zorder=(3 if is_strategy else 2),
            )
        ax.set_xticks(list(x))
        ax.set_xticklabels(horizons_labels)
        ax.set_title(title)
        ax.axhline(0, color="gray", linewidth=0.5)
        if suffix:
            ax.set_ylabel(suffix)
    axes[0, 0].legend(loc="upper left", fontsize=7)
    plt.tight_layout()
    out_png = OUTPUTS_DIR / "benchmark_comparison_chart.png"

    def _save_png(p: Path) -> None:
        fig.savefig(p, dpi=150)

    try:
        retry_on_permission_error(
            lambda: _save_png(out_png), max_attempts=4, base_wait=0.5, logger=logger
        )
        written_png = out_png
    except PermissionError:
        written_png = (
            OUTPUTS_DIR
            / f"benchmark_comparison_chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        )
        logger.warning(
            "Could not write %s after retries (file may be open). Writing %s",
            out_png.name,
            written_png.name,
        )
        _save_png(written_png)
    plt.close()
    logger.info("Wrote %s", written_png)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
