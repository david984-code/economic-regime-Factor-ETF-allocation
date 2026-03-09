"""Excel export for optimal allocations."""

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from src.config import OUTPUTS_DIR

logger = logging.getLogger(__name__)

THRESHOLD = 0.0001
PCT_DECIMALS = 2


def format_allocations_to_excel() -> Path:
    """Format optimal_allocations.csv to Excel. Returns path to output file."""
    input_csv = OUTPUTS_DIR / "optimal_allocations.csv"
    output_xlsx = OUTPUTS_DIR / "optimal_allocations_formatted.xlsx"

    if not input_csv.exists():
        raise FileNotFoundError(f"Missing {input_csv}. Run optimizer first.")

    df = pd.read_csv(input_csv)
    if "regime" in df.columns:
        df = df.set_index("regime")
    else:
        df = df.set_index(df.columns[0])

    df = df.astype(float)
    df[np.abs(df) < THRESHOLD] = 0.0

    row_sums = df.sum(axis=1)
    for idx, s in row_sums.items():
        if s > 0:
            df.loc[str(idx)] = df.loc[str(idx)] / s

    df = df.round(6)
    row_sums = df.sum(axis=1)
    for idx, s in row_sums.items():
        if s > 0:
            df.loc[str(idx)] = df.loc[str(idx)] / s

    df_out = df.copy()
    df_out.insert(0, "regime", df_out.index)

    wb = Workbook()
    ws = wb.active
    ws.title = "allocations"
    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(horizontal="left")

    pct_fmt = "0." + ("0" * PCT_DECIMALS) + "%"
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float, np.floating)):
                cell.number_format = pct_fmt

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    wb.save(output_xlsx)
    logger.info("Saved Excel: %s", output_xlsx)
    return output_xlsx


def main() -> None:
    """Entry point."""
    format_allocations_to_excel()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
