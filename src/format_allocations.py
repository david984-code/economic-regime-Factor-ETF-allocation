from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# ======================
# PATHS
# ======================
ROOT_DIR = Path(__file__).resolve().parents[1]
OUTPUTS_DIR = ROOT_DIR / "outputs"
OUTPUTS_DIR.mkdir(exist_ok=True)

INPUT_CSV = OUTPUTS_DIR / "optimal_allocations.csv"
OUTPUT_XLSX = OUTPUTS_DIR / "optimal_allocations_formatted.xlsx"

# ======================
# SETTINGS
# ======================
THRESHOLD = 0.0001  # 0.01% cutoff
PCT_DECIMALS = 2  # percentage decimals


def main():
    # ----------------------
    # Load data
    # ----------------------
    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"Missing {INPUT_CSV}. Run `python -m src.optimizer` first.")

    df = pd.read_csv(INPUT_CSV)

    # Expect first column to be regime
    if "regime" in df.columns:
        df = df.set_index("regime")
    else:
        df = df.set_index(df.columns[0])

    # ----------------------
    # Clean & normalize
    # ----------------------
    df = df.astype(float)
    df[np.abs(df) < THRESHOLD] = 0.0

    # Normalize rows to sum to 1
    row_sums = df.sum(axis=1)
    for idx, s in row_sums.items():
        if s > 0:
            df.loc[idx] = df.loc[idx] / s

    # Optional rounding + renormalize
    df = df.round(6)
    row_sums = df.sum(axis=1)
    for idx, s in row_sums.items():
        if s > 0:
            df.loc[idx] = df.loc[idx] / s

    # Add regime back as first column
    df_out = df.copy()
    df_out.insert(0, "regime", df_out.index)

    # ----------------------
    # Build Excel workbook
    # ----------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "allocations"

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    bold = Font(bold=True)

    # Bold header row
    for cell in ws[1]:
        cell.font = bold

    # Bold first column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(horizontal="left")

    # Percentage formatting
    percent_format = "0." + ("0" * PCT_DECIMALS) + "%"
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float, np.floating)):
                cell.number_format = percent_format

    # Autosize columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    wb.save(OUTPUT_XLSX)
    print(f"[SUCCESS] Saved formatted Excel file: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
